from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT.parent / "01_Data"
OUT = ROOT / "public" / "dashboard-data.json"
IMPORT_REPORT = ROOT / "public" / "data-import-report.json"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Stock product groups are defined here once and emitted with every stock row.
# Client calculations consume productGroup instead of reclassifying source labels.
STOCK_PRODUCT_CONFIG = {
    "UNIT_PRODUCTS": ("TT", "CH", "EX", "TP", "MAX"),
    "VALUE_ONLY_PRODUCTS": ("IM", "IMO", "OT"),
    "ALL_VALUE_PRODUCTS": ("TT", "CH", "EX", "TP", "MAX", "IM", "IMO", "OT"),
}


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def identifier_key(value: Any) -> str:
    return re.sub(r"[\u200B-\u200D\uFEFF]", "", as_text(value).upper())


def as_number(value: Any) -> float:
    if value is None or value == "":
        return 0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0


def optional_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        text = str(value).strip().replace(",", "")
        if text.startswith("(") and text.endswith(")"):
            text = f"-{text[1:-1]}"
        cleaned = re.sub(r"[^0-9.\-]", "", text)
        return float(cleaned) if cleaned not in {"", "-", "."} else None
    except ValueError:
        return None


def as_int(value: Any) -> int:
    return int(as_number(value))


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = as_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def year_month_from_code(value: Any) -> tuple[int | None, int | None]:
    text = as_text(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 4:
        return None, None
    year = 2000 + int(digits[:2])
    month = int(digits[2:4])
    if month < 1 or month > 12:
        return year, None
    return year, month


def year_month_from_date(value: Any) -> tuple[int | None, int | None]:
    if isinstance(value, datetime):
        return value.year, value.month
    text = iso_date(value)
    if len(text) >= 7 and text[:4].isdigit() and text[5:7].isdigit():
        return int(text[:4]), int(text[5:7])
    return None, None


def normalize_branch(value: Any) -> str:
    text = as_text(value).upper().replace(" ", "")
    if text in {"KMM1", "KMM01"}:
        return "KMM01"
    if text in {"KMM2", "KMM02"}:
        return "KMM02"
    if text in {"KMM3", "KMM03"}:
        return "KMM03"
    return text


def normalize_product_type(value: Any) -> str:
    """Return the approved product group represented by a booking source value."""
    raw = as_text(value).upper()
    key = re.sub(r"[^A-Z0-9]", "", raw)
    aliases = {
        "TT": {"TT", "01TT", "TRACTOR", "01TRACTOR"},
        "CH": {"CH", "02CH", "COMBINE", "COMBINEHARVESTER", "02COMBINE", "02COMBINEHARVESTER"},
        "EX": {"EX", "04EX", "EXCAVATOR", "04EXCAVATOR"},
        "TP": {"TP", "03TP", "TRANSPLANTER", "03TRANSPLANTER"},
        "MAX": {"MAX", "05MAX", "MAXOTHER", "05MAXOTHER"},
        "IM": {"IM", "IMPLEMENT", "IMPLEMENTS"},
        "IMO": {"IMO", "IMPLEMENTOPTION", "IMPLEMENTOPTIONS"},
        "OT": {"OT", "OTHER", "OTHERS"},
    }
    for product, values in aliases.items():
        if key in values:
            return product
    return raw


def normalize_purchase_status(value: Any) -> str:
    return " ".join(as_text(value).upper().split())


def is_hot_booking(row: dict[str, Any]) -> bool:
    return row["purchaseStatus"] in {"A HOT", "B HOT", "C HOT"}


TRACTOR_MODEL_PATTERNS = (
    re.compile(r"^NSPU[A-Z0-9-]*$"),
    re.compile(r"^MU[0-9][A-Z0-9-]*$"),
    re.compile(r"^M[0-9][A-Z0-9-]*$"),
    re.compile(r"^L[0-9][A-Z0-9-]*$"),
    re.compile(r"^B[0-9][A-Z0-9-]*$"),
)


def normalize_stock_model_for_fallback(value: Any) -> str:
    model = " ".join(as_text(value).upper().split())
    model = re.sub(r"\s*\((DEMO|SECOND)\)\s*$", "", model)
    model = re.sub(r"(?:\+(?:FD|FG))+$", "", model)
    return model.strip()


def classify_stock_model_fallback(value: Any) -> str:
    model = normalize_stock_model_for_fallback(value)
    return "TT" if any(pattern.match(model) for pattern in TRACTOR_MODEL_PATTERNS) else "Unknown"


def normalize_stock_product_type(value: Any, model: Any = None) -> str:
    """Normalize Stock TYPE values without silently assigning unknown source labels."""
    raw = " ".join(as_text(value).upper().split())
    key = re.sub(r"[^A-Z0-9]", "", raw)
    aliases = {
        "TT": {"TT", "01TT"},
        "CH": {"CH", "02CH"},
        "EX": {"EX", "03EX"},
        "TP": {"TP", "04TP"},
        "MAX": {"MAX", "05MAX"},
        "IM": {"IM", "06IM", "IMPLEMENT", "IMPLEMENTS"},
        "IMO": {"IMO", "07IMO", "IMPLEMENTOTHER", "IMPLEMENTOTHERS"},
        "OT": {"OT", "OTHER", "OTHERS"},
    }
    for group, values in aliases.items():
        if key in values:
            return group
    return classify_stock_model_fallback(model)


def normalize_kmm_value(value: Any) -> int | None:
    """Return the canonical KMM ownership flag; only 1 qualifies as KMM stock."""
    text = as_text(value)
    return 1 if text == "1" else None


def normalize_status_pd(value: Any) -> str:
    return " ".join(as_text(value).casefold().split())


def is_kmm_free_stock(kmm: Any, status_pd: Any) -> bool:
    return normalize_kmm_value(kmm) == 1 and normalize_status_pd(status_pd) == "free stock"


def physical_machine_key(row: dict[str, Any]) -> str | None:
    for label, field in (("chassis", "chassisNumber"), ("engine", "engineNumber"), ("serial", "serialNumber"), ("stock", "stockId")):
        normalized = identifier_key(row.get(field))
        if normalized and normalized != "-":
            return f"{label}:{normalized}"
    return None


def header_map(ws: Any, row: int) -> dict[str, int]:
    return {as_text(ws.cell(row, col).value): col for col in range(1, ws.max_column + 1) if as_text(ws.cell(row, col).value)}


def cell(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    col = headers.get(name)
    if not col or col - 1 >= len(row):
        return None
    return row[col - 1]


def load_plan() -> dict[str, Any]:
    path = DATA_ROOT / "Sales" / "Sales KPI.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["SL 2026"]
    categories = {as_text(ws.cell(row, 2).value): row for row in range(1, ws.max_row + 1)}
    month_cols = list(range(4, 16))

    def row_values(label: str) -> list[float]:
        row = categories[label]
        return [as_number(ws.cell(row, col).value) for col in month_cols]

    return {
        "year": 2026,
        "months": MONTHS,
        "units": row_values("Total Units"),
        "revenue": row_values("Revenue (2026)"),
        "expense": row_values("Expense (2026)"),
    }


def load_sales() -> list[dict[str, Any]]:
    path = DATA_ROOT / "Sales" / "2026_KMM_CPI copy.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["2026_KMM_DATA"]
    headers = header_map(ws, 4)
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        year = as_int(cell(row, headers, "KMM Year"))
        if not year:
            continue
        code_year, month = year_month_from_code(cell(row, headers, "KMM RS"))
        date = iso_date(cell(row, headers, "Delivery Date"))
        if month is None:
            _, month = year_month_from_date(cell(row, headers, "Delivery Date"))
        rows.append(
            {
                "date": date,
                "year": code_year or year,
                "month": month,
                "branch": normalize_branch(cell(row, headers, "Dealer")),
                "salesperson": as_text(cell(row, headers, "Sales Man")),
                "stateRegion": as_text(cell(row, headers, "States / Division / Region")),
                "township": as_text(cell(row, headers, "Township")),
                "village": as_text(cell(row, headers, "Village")),
                "area": as_text(cell(row, headers, "Area")),
                "productType": as_text(cell(row, headers, "TYPE")),
                "model": as_text(cell(row, headers, "MODEL")),
                "finalReceived": as_number(cell(row, headers, "Final Received")),
                "netReceived": as_number(cell(row, headers, "Net Received")),
                "gp1": as_number(cell(row, headers, "GP1")),
                "commission": as_number(cell(row, headers, "Total")),
                "expense": as_number(cell(row, headers, "Total Expense")),
            }
        )
    return rows


def load_booking() -> list[dict[str, Any]]:
    path = DATA_ROOT / "Booking" / "KMMF3002 KMM Booking Data.2.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data"]
    headers = header_map(ws, 3)
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        booking_no = as_text(cell(row, headers, "No.BK"))
        date = iso_date(cell(row, headers, "Date"))
        branch = normalize_branch(cell(row, headers, "Dealer"))
        salesperson = as_text(cell(row, headers, "SL Name"))
        product_type = normalize_product_type(cell(row, headers, "Product Type"))

        # A valid booking needs an identifier, date, dealer, salesperson and an
        # approved product group. Price is deliberately not a validity rule:
        # source rows with a missing price still represent a booked unit, but do
        # not contribute a monetary amount.
        if not all([booking_no, date, branch, salesperson]) or product_type not in {
            "TT", "CH", "EX", "TP", "MAX", "IM", "IMO", "OT"
        }:
            continue

        year, month = year_month_from_code(cell(row, headers, "Month"))
        month_out = as_text(cell(row, headers, "Month Out")).upper()
        # In this workbook Month Out carries the outcome: CANCLE is cancelled,
        # a numeric month means delivered, and blank remains open. Purchase
        # Status mirrors this (Fail/Sell/Hot) but is not the status field.
        if "CANC" in month_out:
            status = "Cancelled"
        elif month_out:
            status = "Delivered"
        else:
            status = "Open"
        rows.append(
            {
                "date": date,
                "year": year,
                "month": month,
                "branch": branch,
                "salesperson": salesperson,
                "productType": product_type,
                "model": as_text(cell(row, headers, "Model")),
                "price": optional_number(cell(row, headers, "Price")),
                "bookingNo": booking_no,
                "customer": as_text(cell(row, headers, "CS NAME")),
                "deposit": optional_number(cell(row, headers, "Deposit")),
                "paymentType": as_text(cell(row, headers, "Purchase Type")),
                "financeType": as_text(cell(row, headers, "Leasing")),
                "purchaseStatus": normalize_purchase_status(cell(row, headers, "Purchase Status")),
                "statusDate": iso_date(cell(row, headers, "Delivery Dete/Cancer Date")),
                "status": status,
            }
        )
    return rows


def booking_reconciliation_report(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Report the same workbook-backed booking rules used by the dashboard."""
    open_rows = [row for row in rows if is_hot_booking(row)]
    by_product = {
        product: {
            "openRows": sum(row["productType"] == product for row in open_rows),
            "openUnit": sum(row["productType"] == product for row in open_rows),
            "openValue": sum(float(row["price"] or 0) for row in open_rows if row["productType"] == product),
            "deposit": sum(float(row["deposit"] or 0) for row in open_rows if row["productType"] == product),
        }
        for product in ["TT", "CH", "EX", "TP", "MAX", "IM", "IMO", "OT"]
    }
    by_branch = {
        branch: {
            "openUnit": sum(row["branch"] == branch for row in open_rows),
            "openValue": sum(float(row["price"] or 0) for row in open_rows if row["branch"] == branch),
            "deposit": sum(float(row["deposit"] or 0) for row in open_rows if row["branch"] == branch),
        }
        for branch in ["KMM01", "KMM02", "KMM03"]
    }
    return {
        "rawBookingRows": len(rows),
        "cancelledRows": sum(row["status"] == "Cancelled" for row in rows),
        "deliveredRows": sum(row["status"] == "Delivered" for row in rows),
        "purchaseStatusCounts": {status: sum(row["purchaseStatus"] == status for row in rows) for status in ["A HOT", "B HOT", "C HOT"]},
        "openRows": len(open_rows),
        "openUnitTotal": len(open_rows),
        "openValueTotal": sum(float(row["price"] or 0) for row in open_rows),
        "depositTotal": sum(float(row["deposit"] or 0) for row in open_rows),
        "byProduct": by_product,
        "byBranch": by_branch,
    }


def load_stock() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    path = DATA_ROOT / "Stock" / "2026_KMM_R2_STOCK.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["02) STOCK"]
    headers = header_map(ws, 4)
    required_headers = ["STOCK CODE", "BRANCH", "TYPE", "MODEL", "CHASSIS NUMBER", "ENGINE NUMBER", "DAY IN", "Today", "Age Stock", "Car Flow", "KMM", "Status PD", "MSRP", "Sales Staff"]
    missing_headers = [name for name in required_headers if name not in headers]
    if missing_headers:
        raise ValueError(f"Stock import aborted: missing required header(s): {', '.join(missing_headers)}")
    rows: list[dict[str, Any]] = []
    all_rows = list(ws.iter_rows(min_row=5, values_only=True))
    excluded_kmm: Counter[str] = Counter()
    excluded_statuses: Counter[str] = Counter()
    excluded_samples: dict[str, list[dict[str, Any]]] = defaultdict(list)
    kmm_one_rows = 0
    free_stock_rows = 0
    kmm_free_stock_rows = 0
    raw_current: list[tuple[int, tuple[Any, ...]]] = []
    for excel_row, row in enumerate(all_rows, start=5):
        kmm = cell(row, headers, "KMM")
        status_pd = as_text(cell(row, headers, "Status PD"))
        kmm_is_one = normalize_kmm_value(kmm) == 1
        status_is_free = normalize_status_pd(status_pd) == "free stock"
        kmm_one_rows += kmm_is_one
        free_stock_rows += status_is_free
        sample = {"excelRow": excel_row, "stockId": as_text(cell(row, headers, "STOCK CODE")) or None, "serialNumber": as_text(cell(row, headers, "CHASSIS NUMBER")) or None, "engineNumber": as_text(cell(row, headers, "ENGINE NUMBER")) or None, "model": as_text(cell(row, headers, "MODEL")), "productType": as_text(cell(row, headers, "TYPE")), "kmm": as_text(kmm), "statusPd": status_pd, "branch": normalize_branch(cell(row, headers, "BRANCH")) or "Missing"}
        if not kmm_is_one:
            excluded_kmm[as_text(kmm) or "<blank>"] += 1
            if len(excluded_samples["kmmNotOne"]) < 3:
                excluded_samples["kmmNotOne"].append(sample)
            continue
        if not status_is_free:
            excluded_statuses[status_pd or "<blank>"] += 1
            if len(excluded_samples["statusNotFreeStock"]) < 3:
                excluded_samples["statusNotFreeStock"].append(sample)
            continue
        kmm_free_stock_rows += 1
        raw_current.append((excel_row, row))
        branch = normalize_branch(cell(row, headers, "BRANCH")) or "Missing"
        date_in = iso_date(cell(row, headers, "DAY IN"))
        year, month = year_month_from_date(cell(row, headers, "DAY IN"))
        raw_type = as_text(cell(row, headers, "TYPE"))
        model = as_text(cell(row, headers, "MODEL"))
        rows.append(
            {
                "date": date_in,
                "year": year,
                "month": month,
                "branch": branch,
                "salesperson": as_text(cell(row, headers, "Sales Staff")),
                "kmm": normalize_kmm_value(kmm),
                "stockId": as_text(cell(row, headers, "STOCK CODE")) or None,
                "productType": raw_type,
                "productGroup": normalize_stock_product_type(raw_type, model),
                "model": model,
                "ageBucket": as_text(cell(row, headers, "Car Flow")),
                "ageDays": optional_number(cell(row, headers, "Age Stock")),
                "snapshotDate": iso_date(cell(row, headers, "Today")),
                "msrp": optional_number(cell(row, headers, "MSRP")),
                "serialNumber": as_text(cell(row, headers, "CHASSIS NUMBER")) or None,
                "chassisNumber": as_text(cell(row, headers, "CHASSIS NUMBER")) or None,
                "engineNumber": as_text(cell(row, headers, "ENGINE NUMBER")) or None,
                "currentStatus": status_pd,
            }
        )
    known_value = set(STOCK_PRODUCT_CONFIG["ALL_VALUE_PRODUCTS"])
    unit_groups = set(STOCK_PRODUCT_CONFIG["UNIT_PRODUCTS"])
    groups = [*STOCK_PRODUCT_CONFIG["ALL_VALUE_PRODUCTS"], "Unknown"]
    # The source has no quantity field: one current row is one unit unless the
    # physical machine identifier repeats. Stock ID is only a fallback when
    # chassis, engine and serial are blank.
    deduplicated_rows: list[dict[str, Any]] = []
    deduplicated_source_rows: list[tuple[int, tuple[Any, ...]]] = []
    duplicate_records: list[dict[str, Any]] = []
    seen_identifiers: dict[str, int] = {}
    for source_row, imported in zip(raw_current, rows):
        excel_row = source_row[0]
        identifier = physical_machine_key(imported)
        if identifier and identifier in seen_identifiers:
            duplicate_records.append({"identifier": identifier, "keptExcelRow": seen_identifiers[identifier], "removedExcelRow": excel_row, "stockId": imported["stockId"], "serialNumber": imported["serialNumber"], "engineNumber": imported["engineNumber"], "productType": imported["productType"], "branch": imported["branch"], "currentStatus": imported["currentStatus"], "dateIn": imported["date"]})
            continue
        if identifier:
            seen_identifiers[identifier] = excel_row
        deduplicated_rows.append(imported)
        deduplicated_source_rows.append(source_row)
    rows = deduplicated_rows
    rows_by_group = {group: [row for row in rows if row["productGroup"] == group] for group in groups}
    for group, reason in [("IM", "productIM"), ("IMO", "productIMO"), ("OT", "productOT"), ("Unknown", "missingOrUnknownProduct")]:
        for row in rows_by_group[group][:3]:
            excluded_samples[reason].append({"stockId": row["stockId"], "serialNumber": row["serialNumber"], "engineNumber": row["engineNumber"], "model": row["model"], "productType": row["productType"], "kmm": row["kmm"], "statusPd": row["currentStatus"], "branch": row["branch"]})
    for reason in ["kmmNotOne", "statusNotFreeStock", "productIM", "productIMO", "productOT", "missingOrUnknownProduct"]:
        excluded_samples.setdefault(reason, [])
    serial_keys: dict[tuple[str, str], list[int]] = defaultdict(list)
    fallback_keys: dict[tuple[str, str, str, str, str], list[int]] = defaultdict(list)
    missing_serial = 0
    for (excel_row, source_row), imported in zip(deduplicated_source_rows, rows):
        serial = imported["serialNumber"]
        if serial and serial != "-":
            serial_keys[(imported["branch"], serial)].append(excel_row)
        else:
            missing_serial += 1
        fallback_keys[(imported["branch"], imported["productType"], imported["model"], imported["date"], str(imported["msrp"]))].append(excel_row)
    exact_duplicates = {" | ".join(key): locations for key, locations in serial_keys.items() if len(locations) > 1}
    possible_duplicates = {" | ".join(key): locations for key, locations in fallback_keys.items() if len(locations) > 1}
    report = {
        "source": str(path.relative_to(ROOT.parent)),
        "sheet": "02) STOCK",
        "headerRow": 4,
        "headers": [as_text(ws.cell(4, col).value) for col in range(1, ws.max_column + 1) if as_text(ws.cell(4, col).value)],
        "currentStockRule": "KMM = 1 and Status PD = Free Stock (case-insensitive, trimmed) are both required.",
        "quantityRule": "No Quantity column exists; one valid current-stock row equals one unit.",
        "snapshotDate": next((row["snapshotDate"] for row in rows if row["snapshotDate"]), None),
        "excelTotalRows": len(all_rows),
        "totalSourceRows": len(all_rows),
        "kmmOneRows": kmm_one_rows,
        "statusPdFreeStockRows": free_stock_rows,
        "kmmFreeStockRows": kmm_free_stock_rows,
        "currentStockRows": len(rows),
        "validRemainingStockRows": len(rows),
        "excludedRows": len(all_rows) - kmm_free_stock_rows,
        "excludedByKmm": dict(sorted(excluded_kmm.items())),
        "excludedSoldDeliveredRows": sum(excluded_statuses.values()),
        "excludedByStatusPD": dict(sorted(excluded_statuses.items())),
        "excludedSamples": dict(excluded_samples),
        "missingProductType": sum(not row["productType"] for row in rows),
        "unknownProductTypes": sorted({row["productType"] for row in rows if row["productGroup"] not in known_value}),
        "unknownProductRows": sum(row["productGroup"] not in known_value for row in rows),
        "missingBranch": sum(row["branch"] == "Missing" for row in rows),
        "missingModel": sum(not row["model"] for row in rows),
        "missingValue": sum(row["msrp"] is None for row in rows),
        "invalidValue": 0,
        "invalidDateIn": sum(not row["date"] for row in rows),
        "negativeAge": sum(row["ageDays"] is not None and row["ageDays"] < 0 for row in rows),
        "inconsistentAge": sum(bool(row["date"] and row["snapshotDate"] and row["ageDays"] is not None and abs((datetime.fromisoformat(row["snapshotDate"]) - datetime.fromisoformat(row["date"])).days - row["ageDays"]) > 1) for row in rows),
        "rowsWithoutSerialNumber": missing_serial,
        "exactDuplicateRows": len(duplicate_records),
        "exactDuplicateKeys": exact_duplicates,
        "duplicateRecordsRemoved": duplicate_records,
        "possibleDuplicateRows": sum(len(locations) - 1 for locations in possible_duplicates.values()),
        "possibleDuplicateKeys": possible_duplicates,
        "unitTotalByProductType": {group: len(rows_by_group[group]) for group in groups},
        "excludedFromUnitByProductType": {group: len(rows_by_group[group]) for group in ["IM", "IMO", "OT"]},
        "valueTotalByProductType": {group: sum(row["msrp"] or 0 for row in rows_by_group[group]) for group in groups},
        "unitTotalByBranch": {branch: sum(1 for row in rows if row["branch"] == branch and row["productGroup"] in unit_groups) for branch in ["KMM01", "KMM02", "KMM03"]},
        "otherOrUnknownBranchUnit": sum(1 for row in rows if row["branch"] not in {"KMM01", "KMM02", "KMM03"} and row["productGroup"] in unit_groups),
        "stockUnit": sum(1 for row in rows if row["productGroup"] in unit_groups),
        "stockValue": sum(row["msrp"] or 0 for row in rows if row["productGroup"] in known_value),
        "agedStockUnit": sum(1 for row in rows if row["productGroup"] in unit_groups and (row["ageDays"] or 0) > 90),
    }
    return rows, report


def load_marketing() -> list[dict[str, Any]]:
    path = DATA_ROOT / "Marketing" / "2026 Marketing Summary.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = header_map(ws, 2)
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        year, month = year_month_from_date(cell(row, headers, "Event Date"))
        rows.append(
            {
                "date": iso_date(cell(row, headers, "Event Date")),
                "year": year,
                "month": month,
                "branch": normalize_branch(cell(row, headers, "Dealer")),
                "salesperson": "",
                "stateRegion": as_text(cell(row, headers, "Division")),
                "township": as_text(cell(row, headers, "Township")),
                "village": as_text(cell(row, headers, "Village")),
                "activity": as_text(cell(row, headers, "Type of Activities")),
                "participants": as_int(cell(row, headers, "Participants")),
                "bookingCount": as_int(cell(row, headers, "BK")),
                "prospectCount": as_int(cell(row, headers, "PC")),
                "expense": as_number(cell(row, headers, "KMM Expense")),
            }
        )
    return rows


def main() -> None:
    sources = [
        DATA_ROOT / "Sales" / "Sales KPI.xlsx",
        DATA_ROOT / "Sales" / "2026_KMM_CPI copy.xlsx",
        DATA_ROOT / "Booking" / "KMMF3002 KMM Booking Data.2.xlsx",
        DATA_ROOT / "Stock" / "2026_KMM_R2_STOCK.xlsx",
        DATA_ROOT / "Marketing" / "2026 Marketing Summary.xlsx",
    ]
    latest_mtime = max(path.stat().st_mtime for path in sources)
    stock, stock_report = load_stock()
    data = {
        "meta": {
            "company": "KUBOTA MAESOD MYANMAR",
            "shortName": "KMM",
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceUpdatedAt": datetime.fromtimestamp(latest_mtime).isoformat(timespec="seconds"),
            "sources": [str(path.relative_to(ROOT.parent)) for path in sources],
        },
        "plan": load_plan(),
        "sales": load_sales(),
        "booking": load_booking(),
        "stock": stock,
        "marketing": load_marketing(),
    }
    report = {"generatedAt": datetime.now().isoformat(timespec="seconds"), "stock": stock_report, "booking": booking_reconciliation_report(data["booking"])}
    out_tmp = OUT.with_suffix(".tmp")
    report_tmp = IMPORT_REPORT.with_suffix(".tmp")
    out_tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    report_tmp.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    out_tmp.replace(OUT)
    report_tmp.replace(IMPORT_REPORT)
    print(f"Wrote {OUT}")
    print(f"Wrote {IMPORT_REPORT}")
    print({key: len(data[key]) for key in ["sales", "booking", "stock", "marketing"]})
    booking_report = report["booking"]
    print({
        "totalRows": booking_report["rawBookingRows"],
        "aHot": booking_report["purchaseStatusCounts"]["A HOT"],
        "bHot": booking_report["purchaseStatusCounts"]["B HOT"],
        "cHot": booking_report["purchaseStatusCounts"]["C HOT"],
        "openBookingUnit": booking_report["openUnitTotal"],
        "bookingValuePrice": booking_report["openValueTotal"],
        "depositReceived": booking_report["depositTotal"],
        "bookingValueByBranch": booking_report["byBranch"],
        "bookingValueByProduct": booking_report["byProduct"],
    })


if __name__ == "__main__":
    main()
